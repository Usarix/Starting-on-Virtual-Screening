import requests
from bs4 import BeautifulSoup
import xlsxwriter
import openpyxl

import sys


workbook = xlsxwriter.Workbook('ZINCSFromPages(8QPV).xlsx')     #THE NAME OF THE EXCEL FILE YOU WANT YOUR DATA
wb = openpyxl.load_workbook('bestranking_2(8QPV).xlsx')         #EXCEL WITH THE ZINC ID OF THE MOLECULES AND THE SCORE OBTAIN FROM PLANTS ex: ZINC0000123456, -100
ws = wb["bestranking_2"]                                        #NAME OF THE EXCEL PAGE WHERE THE DATA IS
#listademoleculas = [ws.cell(row=i+1,column=3).value for i in range(ws.max_row-10,3)]


#print(ws.cell(ws.max_row - 1,3).value)

#print(ws.max_row)

worksheet = workbook.add_worksheet()

#print(listademoleculas)

#for molecules in listademoleculas:
#    print(molecules)
matrixInformation = [["ID","Score","Ph Range", "Net Charge","H-bond Donors","H-bond Acceptors","tPSA","Rotatable Bonds","Apolar Desolvation","Polar Desolavation","Download"]]

for fff in range(2,ws.max_row - 1):     #2 is the row, normally the first row is for indexes, that is why it starts on 2
        
    
    moleculeID = ws.cell(fff, 3).value      #the 3 is in which column is the IDs. 1 = A, 2 = B...
    Score = ws.cell(fff,6).value            #the 6 is the same, its the column
    # URL de la página web que deseas analizar
    #print(moleculeID)
    url = "https://zinc15.docking.org/substances/"+moleculeID+"/"


    

    # Realizar una solicitud HTTP GET a la página web
    response = requests.get(url)

    # Verificar que la solicitud fue exitosa
    if response.status_code == 200:
        # Analizar el contenido HTML
        soup = BeautifulSoup(response.content, "html.parser")
        
        # Ejemplo: encontrar un bloque específico por su identificador (id) o clase
        bloque = soup.find("div", {"class": "protomers panel panel-default row panel-"})  # Cambia "id" o "class" según corresponda
        
        if bloque:
            # Extraer la información dentro del bloque
            tabla = bloque.find("table", {"class": "table table-striped table-condensed table-hover table-numeric table-responsive table-sm-collapse"})
            if tabla:
                body = tabla.find("tbody")
                if body:
                    rows = body.find_all("tr")
                    info = body.find("tr")
                    if info:

                        indices_ii = 2

                        for indices in rows:
                            
                            fila = [moleculeID, Score]
                            cellinfo = indices.find_all("td")
                            
                            celda_ii = 0
                            
                            for celdas in cellinfo:

                                #print(celdas.get_text())
                                #matrixInformation[indices_ii][celda_ii] = celdas.get_text()
                                textToString = celdas.get_text()
                                
                                #print(textToString)
                                fila.append(textToString.strip())
                                #print(fila)
                                
                                celda_ii = celda_ii + 1
                                #if celda_ii > 8:
                                    #print("it activates")
                            #print(fila)
                            matrixInformation.append(fila)
                            #print(((matrixInformation[-1][0])))
                            #print((len(matrixInformation[-1][0])))
                            if (int(matrixInformation[-1][3]) <= 0):
                                matrixInformation.pop(-1)
                                #matrixInformation[indices_ii] = [0,0,0,0,0,0,0,0,0]
                                #indices_ii = indices_ii
                            
                            #elif (len(matrixInformation[-1][0]) != 9) and (len(matrixInformation[-1][0]) != 12):
                            #    matrixInformation.pop(-1)
                            elif (matrixInformation[-1][2] != "Reference") and (matrixInformation[-1][2] != "Mid pH (7.4)"):
                                matrixInformation.pop(-1)
                            #    print(type(matrixInformation[-1][0]))
                                
                            #    matrixInformation.pop(-1)
                                
                                #indices_ii = indices_ii

                            #elif (matrixInformation[indices_ii][0] != "Reference") and (matrixInformation[indices_ii][0] != "Mid pH (7.4)"):
                            #    matrixInformation[indices_ii] = [0,0,0,0,0,0,0,0,0]
                            #    indices_ii = indices_ii
                            
                            
                                #print((matrixInformation[-1][0]))
                                #indices_ii = indices_ii + 1
                            
                            #if indices_ii > 4:
                            #    sys.exit("Tiene mas de dos ph's" + moleculeID)


                            
                        #print(info.find("td", {"title": "Hydrogen Bond Donors"}))
                        #if compossinfo:
                        #    contenido_dentro_del_bloque = compossinfo.get_text()
                        #    print(contenido_dentro_del_bloque)
                        #else:
                        #    print("fail composs")
                        
                    else:
                        print("fallo el info")
                else:
                    print("no hay todo")
                
            else:
                print("tabla no encontrada")
            
            
        else:
            print("No se encontró el bloque específico.")
    else:
        print(f"Error al acceder a la página. Código de estado: {response.status_code}")
        print(moleculeID)

    #worksheet.write(0,0, matrixInformation[0][0])


row = 0
col = 0
#print(matrixInformation[-1])
for identificacion, puntaje, Ph, Charge,Donors,Acceptors,tPSA,Rotatable,Apolar,Polar, Download in matrixInformation:
    worksheet.write(row, col, identificacion)
    worksheet.write(row, col + 1, puntaje)
    worksheet.write(row, col + 2, Ph)
    worksheet.write(row, col + 3, Charge)
    worksheet.write(row, col + 4 , Donors)
    worksheet.write(row, col + 5 , Acceptors)
    worksheet.write(row, col + 6, tPSA)
    worksheet.write(row, col + 7, Rotatable)
    worksheet.write(row, col + 8, Apolar)
    worksheet.write(row, col + 9, Polar)
    #worksheet.write(row,col + 8, Download)
    row += 1
workbook.close()


#print(matrixInformation)

